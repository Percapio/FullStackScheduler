"""Tests for the SCHD sheet layout introduced in Phase 12.

Covers:
- resolve_sheet: all five resolution branches
- resolve_layout: happy path + unknown sheet
- read_rows with SCHD layout: title skip, divider skip, whitespace divider
  treated as data, all-empty row skip, missing-DIVIDER-column error,
  empty-sheet error, None header column skip, eager-validation semantics
- read_rows with SHIPPED (AA) layout: regression (no behavior change)
- Integration: ingest_workbook via SCHD workbook; fallback via SHIPPED-only
"""
from __future__ import annotations

import pytest
from openpyxl import Workbook

from backend.app.reader import (
    FALLBACK_SHEET_NAME,
    PRIMARY_SHEET_NAME,
    EmptySheetError,
    MissingDividerColumnError,
    SheetLayout,
    _LAYOUTS,
    read_rows,
    resolve_layout,
    resolve_sheet,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_wb(*sheet_names: str) -> Workbook:
    """Return an openpyxl Workbook containing sheets with the given names."""
    wb = Workbook()
    # The default sheet is named "Sheet" by openpyxl; rename or add as needed.
    first, *rest = sheet_names
    wb.active.title = first
    for name in rest:
        wb.create_sheet(name)
    return wb


# ---------------------------------------------------------------------------
# resolve_sheet — five branches
# ---------------------------------------------------------------------------


def test_resolve_sheet_prefers_schd():
    """When both SCHD and SHIPPED (AA) are present, SCHD is returned."""
    wb = _make_wb("SCHD", "SHIPPED (AA)")
    assert resolve_sheet(wb, "SCHD") == "SCHD"


def test_resolve_sheet_falls_back_to_shipped():
    """SCHD absent + SHIPPED (AA) present + requested=PRIMARY → fallback."""
    wb = _make_wb("SHIPPED (AA)")
    assert resolve_sheet(wb, PRIMARY_SHEET_NAME) == FALLBACK_SHEET_NAME


def test_resolve_sheet_explicit_no_fallback():
    """Explicitly requesting 'SHIPPED (AA)' against a SCHD-only workbook raises."""
    wb = _make_wb("SCHD")
    with pytest.raises(KeyError):
        resolve_sheet(wb, "SHIPPED (AA)")


def test_resolve_sheet_neither_present():
    """Neither PRIMARY nor FALLBACK present raises KeyError."""
    wb = _make_wb("Sheet1")
    with pytest.raises(KeyError):
        resolve_sheet(wb, PRIMARY_SHEET_NAME)


def test_resolve_sheet_returns_requested_when_present():
    """Requested sheet found verbatim — even when it is SHIPPED (AA) explicitly."""
    wb = _make_wb("SHIPPED (AA)", "SCHD")
    assert resolve_sheet(wb, "SHIPPED (AA)") == "SHIPPED (AA)"


# ---------------------------------------------------------------------------
# resolve_layout
# ---------------------------------------------------------------------------


def test_resolve_layout_schd():
    layout = resolve_layout("SCHD")
    assert isinstance(layout, SheetLayout)
    assert layout.header_row == 2
    assert layout.divider_header == "DIVIDER"


def test_resolve_layout_shipped_aa():
    layout = resolve_layout("SHIPPED (AA)")
    assert isinstance(layout, SheetLayout)
    assert layout.header_row == 1
    assert layout.divider_header is None


def test_resolve_layout_unknown_sheet():
    with pytest.raises(KeyError):
        resolve_layout("Sheet1")


# ---------------------------------------------------------------------------
# SCHD read_rows — layout traversal
# ---------------------------------------------------------------------------


def test_read_rows_skips_title_row(schd_workbook_factory):
    """Title row (row 1) is skipped; first data row is row 3."""
    path = schd_workbook_factory([{"data": {"JOB": "100001\nNEW"}}])
    rows = list(read_rows(str(path), sheet="SCHD"))
    assert len(rows) == 1
    row_number, cells = rows[0]
    assert row_number == 3


def test_read_rows_skips_divider_rows(schd_workbook_factory):
    """Divider rows are dropped; data row numbers reflect their worksheet position."""
    entries = [
        {"data": {"JOB": "100001\nNEW"}},
        {"data": {}, "divider": True},
        {"data": {"JOB": "100002\nNEW"}},
        {"data": {}, "divider": True},
        {"data": {"JOB": "100003\nNEW"}},
    ]
    path = schd_workbook_factory(entries)
    rows = list(read_rows(str(path), sheet="SCHD"))
    assert len(rows) == 3
    numbers = [r[0] for r in rows]
    assert numbers == [3, 5, 7]


def test_read_rows_divider_takes_precedence(schd_workbook_factory):
    """A row with DIVIDER sentinel set is skipped even if JOB cell has data."""
    entries = [{"data": {"JOB": "128764\nNEW"}, "divider": True}]
    path = schd_workbook_factory(entries)
    rows = list(read_rows(str(path), sheet="SCHD"))
    assert rows == []


def test_read_rows_divider_ignores_whitespace(schd_workbook_factory, tmp_path):
    """A DIVIDER cell containing only whitespace is NOT a divider (Decision 5)."""
    from openpyxl import Workbook as OxWorkbook
    from backend.app.reader import KNOWN_HEADERS

    headers = sorted(KNOWN_HEADERS) + ["DIVIDER"]
    wb = OxWorkbook()
    ws = wb.active
    ws.title = "SCHD"
    ws.append(["Production Schedule"])  # row 1
    ws.append(headers)                  # row 2
    # row 3: DIVIDER cell is a single space (whitespace-only)
    row_values = [None] * len(headers)
    job_idx = headers.index("JOB")
    divider_idx = headers.index("DIVIDER")
    row_values[job_idx] = "100005\nNEW"
    row_values[divider_idx] = " "
    ws.append(row_values)
    path = tmp_path / "schd_ws.xlsx"
    wb.save(path)

    rows = list(read_rows(str(path), sheet="SCHD"))
    assert len(rows) == 1
    _, cells = rows[0]
    assert cells.get("JOB") == "100005\nNEW"


def test_read_rows_empty_data_rows_still_filtered(schd_workbook_factory):
    """An all-None data row between sections is skipped (existing all-empty filter)."""
    entries = [
        {"data": {"JOB": "100001\nNEW"}},
        {"data": {}},  # all None — no JOB, nothing
        {"data": {"JOB": "100002\nNEW"}},
    ]
    path = schd_workbook_factory(entries)
    rows = list(read_rows(str(path), sheet="SCHD"))
    assert len(rows) == 2


def test_read_rows_missing_divider_column_errors(tmp_path):
    """SCHD workbook with no DIVIDER column raises MissingDividerColumnError on call."""
    from openpyxl import Workbook as OxWorkbook
    from backend.app.reader import KNOWN_HEADERS

    wb = OxWorkbook()
    ws = wb.active
    ws.title = "SCHD"
    ws.append(["Title"])
    ws.append(sorted(KNOWN_HEADERS))  # no DIVIDER column
    path = tmp_path / "no_divider.xlsx"
    wb.save(path)

    with pytest.raises(MissingDividerColumnError):
        read_rows(str(path), sheet="SCHD")


def test_read_rows_empty_sheet_errors(tmp_path):
    """SCHD workbook with only a title row raises EmptySheetError on call."""
    from openpyxl import Workbook as OxWorkbook

    wb = OxWorkbook()
    ws = wb.active
    ws.title = "SCHD"
    ws.append(["Title"])  # row 1 only; header_row == 2 is missing
    path = tmp_path / "empty_schd.xlsx"
    wb.save(path)

    with pytest.raises(EmptySheetError):
        read_rows(str(path), sheet="SCHD")


def test_read_rows_skips_none_header_column(tmp_path):
    """A None-valued header cell is silently skipped; surrounding columns extracted."""
    from openpyxl import Workbook as OxWorkbook
    from backend.app.reader import KNOWN_HEADERS

    known = sorted(KNOWN_HEADERS)
    # Insert a None between the first two real headers
    headers_with_gap = [known[0], None, known[1]] + known[2:] + ["DIVIDER"]

    wb = OxWorkbook()
    ws = wb.active
    ws.title = "SCHD"
    ws.append(["Production Schedule"])
    ws.append(headers_with_gap)
    # Row 3: first and second real-header positions filled; None column ignored
    row_values = [None] * len(headers_with_gap)
    row_values[0] = "some value"        # known[0] — any KNOWN_HEADERS value
    row_values[2] = "100001\nNEW"       # known[1] — JOB or whatever known[1] is
    ws.append(row_values)
    path = tmp_path / "gap_headers.xlsx"
    wb.save(path)

    rows = list(read_rows(str(path), sheet="SCHD"))
    assert len(rows) == 1
    _, cells = rows[0]
    # The None column must not appear in cells
    assert None not in cells


def test_read_rows_errors_on_call_not_on_iter(tmp_path):
    """MissingDividerColumnError raises on the read_rows() call, not on next()."""
    from openpyxl import Workbook as OxWorkbook
    from backend.app.reader import KNOWN_HEADERS

    wb = OxWorkbook()
    ws = wb.active
    ws.title = "SCHD"
    ws.append(["Title"])
    ws.append(sorted(KNOWN_HEADERS))
    path = tmp_path / "eager.xlsx"
    wb.save(path)

    with pytest.raises(MissingDividerColumnError):
        # Error must be raised here — before any next() call.
        read_rows(str(path), sheet="SCHD")


# ---------------------------------------------------------------------------
# SHIPPED (AA) regression
# ---------------------------------------------------------------------------


def test_read_rows_shipped_aa_unchanged(workbook_factory):
    """SHIPPED (AA) workbook yields the same row stream as pre-Phase-12."""
    path = workbook_factory([{"JOB": "137845\nNEW", "QTY": "10"}])
    rows = list(read_rows(str(path)))
    assert len(rows) == 1
    row_number, cells = rows[0]
    # SHIPPED (AA) has header_row=1 so first data row is row 2
    assert row_number == 2
    assert cells["JOB"] == "137845\nNEW"


# ---------------------------------------------------------------------------
# Integration: ingest_workbook via SCHD path
# ---------------------------------------------------------------------------


def test_ingest_workbook_consumes_schd(schd_workbook_factory, session_factory):
    """SCHD workbook with one valid JOB row → rows_inserted == 1."""
    from sqlalchemy import select

    from backend.app.ingest import ingest_workbook
    from backend.app.models import Job

    path = schd_workbook_factory([{"data": {"JOB": "137845\nNEW", "QTY": "10", "CUSTOMER": "ACME"}}])
    result = ingest_workbook(str(path), session_factory=session_factory)

    assert result.rows_inserted == 1

    with session_factory() as s:
        jobs = s.scalars(select(Job)).all()
    assert len(jobs) == 1


def test_ingest_workbook_falls_back_to_shipped(workbook_factory, session_factory):
    """SHIPPED-only workbook resolves transparently via fallback; rows ingested."""
    from backend.app.ingest import ingest_workbook

    path = workbook_factory([{"JOB": "137845\nNEW", "QTY": "10", "CUSTOMER": "ACME"}])
    result = ingest_workbook(str(path), session_factory=session_factory)

    assert result.rows_inserted == 1


# ---------------------------------------------------------------------------
# Header normalization through read_rows (Phase 22 Part 1)
# ---------------------------------------------------------------------------

import logging

from backend.app.reader import KNOWN_HEADERS


def _schd_with_headers(tmp_path, headers, data_rows, filename="schd_headers.xlsx"):
    """Build a SCHD workbook whose header row is exactly `headers`.

    Pre:  headers is the literal row-2 sequence, de-normalized entries included.
          data_rows is a list of sequences aligned positionally with headers.
    Post: returns the path to the saved workbook.
    """
    from openpyxl import Workbook as OxWorkbook

    wb = OxWorkbook()
    ws = wb.active
    ws.title = "SCHD"
    ws.append(["Production Schedule"])
    ws.append(list(headers))
    for values in data_rows:
        ws.append(list(values))
    path = tmp_path / filename
    wb.save(path)
    return path


def test_read_rows_matches_kit_rel_with_trailing_space(tmp_path):
    """The Phase 22 bug: 'KIT REL ' is bound and its data survives."""
    headers = ["JOB", "KIT REL ", "DIVIDER"]
    path = _schd_with_headers(tmp_path, headers, [["100001\nNEW", "4/17", None]])

    rows = list(read_rows(str(path), sheet="SCHD"))

    assert len(rows) == 1
    _, cells = rows[0]
    assert cells["KIT REL"] == "4/17"


def test_read_rows_tolerates_none_header_cells(tmp_path):
    """ADV ALL IN ONE SCHEDULE carries None at header indices 26 and 29."""
    headers = ["JOB", None, "QTY", None, "DIVIDER"]
    path = _schd_with_headers(tmp_path, headers, [["100001\nNEW", "x", "10", "y", None]])

    rows = list(read_rows(str(path), sheet="SCHD"))

    assert len(rows) == 1
    _, cells = rows[0]
    assert cells == {"JOB": "100001\nNEW", "QTY": "10"}


def test_read_rows_tolerates_non_string_header_cells(tmp_path):
    """int / datetime header cells normalize to None; behaviour is pre-patch."""
    from datetime import datetime as _dt

    headers = ["JOB", 4, _dt(2026, 8, 28), "DIVIDER"]
    path = _schd_with_headers(tmp_path, headers, [["100001\nNEW", "a", "b", None]])

    rows = list(read_rows(str(path), sheet="SCHD"))

    _, cells = rows[0]
    assert cells == {"JOB": "100001\nNEW"}


def test_read_rows_whitespace_only_header_contributes_nothing(tmp_path):
    """A '   ' header cell becomes None — never an empty-string key."""
    headers = ["JOB", "   ", "DIVIDER"]
    path = _schd_with_headers(tmp_path, headers, [["100001\nNEW", "orphan", None]])

    rows = list(read_rows(str(path), sheet="SCHD"))

    _, cells = rows[0]
    assert "" not in cells
    assert cells == {"JOB": "100001\nNEW"}


def test_read_rows_divider_header_with_trailing_space_still_detected(tmp_path):
    """Normalization precedes _resolve_divider_index, so 'DIVIDER ' still resolves."""
    headers = ["JOB", "DIVIDER "]
    path = _schd_with_headers(
        tmp_path, headers, [["100001\nNEW", None], ["100002\nNEW", "X"]]
    )

    rows = list(read_rows(str(path), sheet="SCHD"))

    assert [r[0] for r in rows] == [3]


def test_read_rows_divider_never_reaches_extracted_cells(tmp_path):
    headers = ["JOB", "DIVIDER"]
    path = _schd_with_headers(tmp_path, headers, [["100001\nNEW", None]])

    _, cells = list(read_rows(str(path), sheet="SCHD"))[0]

    assert "DIVIDER" not in cells


def test_read_rows_warns_on_missing_known_header(tmp_path, caplog):
    headers = ["JOB", "DIVIDER"]
    path = _schd_with_headers(tmp_path, headers, [["100001\nNEW", None]])

    with caplog.at_level(logging.WARNING, logger="backend.app.reader"):
        rows = list(read_rows(str(path), sheet="SCHD"))

    assert len(rows) == 1  # the iterator still yields
    warnings = [r for r in caplog.records if r.levelno == logging.WARNING]
    assert any("reader.headers.missing" in r.getMessage() for r in warnings)
    assert any("KIT REL" in r.getMessage() for r in warnings)


def test_read_rows_collision_first_index_wins_and_warns(tmp_path, caplog):
    """Two source columns normalizing onto one header: first wins, WARNING logged."""
    headers = ["JOB", "KIT REL", "KIT REL ", "DIVIDER"]
    path = _schd_with_headers(
        tmp_path, headers, [["100001\nNEW", "kept", "ignored", None]]
    )

    with caplog.at_level(logging.WARNING, logger="backend.app.reader"):
        rows = list(read_rows(str(path), sheet="SCHD"))

    _, cells = rows[0]
    assert cells["KIT REL"] == "kept"
    collision_messages = [
        r.getMessage() for r in caplog.records
        if "reader.headers.collision" in r.getMessage()
    ]
    assert len(collision_messages) == 1
    assert "kept_index=1" in collision_messages[0]
    assert "ignored_index=2" in collision_messages[0]


def test_read_rows_logs_unrecognised_once_per_call_not_per_row(tmp_path, caplog):
    headers = sorted(KNOWN_HEADERS) + ["FEEDER S/U", "DIVIDER"]
    data_rows = [
        [None] * len(headers) for _ in range(3)
    ]
    job_index = headers.index("JOB")
    for index, values in enumerate(data_rows):
        values[job_index] = f"10000{index}\nNEW"
    path = _schd_with_headers(tmp_path, headers, data_rows)

    with caplog.at_level(logging.INFO, logger="backend.app.reader"):
        rows = list(read_rows(str(path), sheet="SCHD"))

    assert len(rows) == 3
    unrecognised_records = [
        r for r in caplog.records if "reader.headers.unrecognised" in r.getMessage()
    ]
    assert len(unrecognised_records) == 1
    assert "FEEDER S/U" in unrecognised_records[0].getMessage()
    assert "DIVIDER" not in unrecognised_records[0].getMessage()


def test_ingest_persists_kit_rel_from_denormalized_header(tmp_path, session_factory):
    """End-to-end: 'KIT REL ' reaches ImportStagingRow.raw_kit_rel."""
    from sqlalchemy import select as _select

    from backend.app.ingest import ingest_workbook
    from backend.app.models import Assembly, ImportStagingRow

    with session_factory() as seed:
        seed.add(Assembly(part_number="100001"))
        seed.commit()

    headers = sorted(KNOWN_HEADERS) + ["DIVIDER"]
    headers[headers.index("KIT REL")] = "KIT REL "
    values = [None] * len(headers)
    values[headers.index("JOB")] = "100001\nNEW"
    values[headers.index("QTY")] = "10"
    values[headers.index("CUSTOMER")] = "ACME"
    values[headers.index("KIT REL ")] = "4/17"
    path = _schd_with_headers(tmp_path, headers, [values], filename="kit_rel.xlsx")

    ingest_workbook(path, session_factory=session_factory)

    with session_factory() as check:
        staged = check.scalars(_select(ImportStagingRow)).all()
        assert len(staged) == 1
        assert staged[0].raw_kit_rel == "4/17"
