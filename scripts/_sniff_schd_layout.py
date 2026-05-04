"""Preflight sniff for the SCHD workbook layout.

Run:
    python -m scripts._sniff_schd_layout <path-to-production-SCHD.xlsx>

Exits 0 on a successful workbook read. Shape anomalies are reported as
fields in the output, not as non-zero exit codes. The operator inspects
the report and confirms the seven gates listed in §2.3 of the Phase 12 TDD.

Keep this script under scripts/ post-deploy: it doubles as a regression-
detection tool whenever the production SCHD workbook is revised.
"""
from __future__ import annotations

import sys
from dataclasses import dataclass, field, fields
from pathlib import Path

from openpyxl import load_workbook

# Pull KNOWN_HEADERS from the canonical source so the sniff matches exactly
# what the pipeline will look for.
from backend.app.reader import KNOWN_HEADERS


@dataclass(frozen=True)
class SchdSniffReport:
    """Snapshot of the SCHD sheet's structural shape.

    Negative-path defaults let sheet_present=False short-circuit cleanly
    without enumerating every field at the call site.
    """

    sheet_present: bool
    title_row_first_cell: str | None = None
    header_row_values: tuple[str | None, ...] = ()
    divider_column_index: int | None = None
    divider_row_numbers: tuple[int, ...] = ()
    missing_known_headers: tuple[str, ...] = ()  # KNOWN_HEADERS − non-null headers
    extra_headers: tuple[str, ...] = ()           # non-null headers − KNOWN_HEADERS − {"DIVIDER"}


def sniff_schd(path: Path) -> SchdSniffReport:
    """Return a structural report for the SCHD sheet in *path*.

    Pre:  path is a readable .xlsx workbook.
    Post: returns a SchdSniffReport. Never mutates the file.
    Raises: FileNotFoundError when path does not exist; openpyxl load errors
            on a corrupt or non-xlsx file.
    """
    wb = load_workbook(path, data_only=True)
    if "SCHD" not in wb.sheetnames:
        return SchdSniffReport(sheet_present=False)

    ws = wb["SCHD"]
    title_row = next(ws.iter_rows(min_row=1, max_row=1))
    header_row = next(ws.iter_rows(min_row=2, max_row=2))
    headers = tuple(c.value for c in header_row)
    non_null = set(filter(None, headers))

    divider_idx: int | None = None
    if "DIVIDER" in headers:
        divider_idx = list(headers).index("DIVIDER")

    divider_rows: list[int] = []
    if divider_idx is not None:
        for row_number, row in enumerate(ws.iter_rows(min_row=3), start=3):
            if divider_idx < len(row) and row[divider_idx].value not in (None, ""):
                divider_rows.append(row_number)

    return SchdSniffReport(
        sheet_present=True,
        title_row_first_cell=title_row[0].value,
        header_row_values=headers,
        divider_column_index=divider_idx,
        divider_row_numbers=tuple(divider_rows),
        missing_known_headers=tuple(sorted(set(KNOWN_HEADERS) - non_null)),
        extra_headers=tuple(sorted(non_null - set(KNOWN_HEADERS) - {"DIVIDER"})),
    )


def _print_report(report: SchdSniffReport) -> None:
    print("=" * 60)
    print("SCHD Layout Sniff Report")
    print("=" * 60)
    for f in fields(report):
        print(f"  {f.name}: {getattr(report, f.name)!r}")
    print()
    if not report.sheet_present:
        print("GATE 1 FAIL: sheet_present=False — SCHD tab not found in workbook.")
        return

    print("Gate results (operator must confirm each):")
    gates = [
        ("1. sheet_present=True",                   report.sheet_present),
        ("2. title_row_first_cell is non-null",      report.title_row_first_cell is not None),
        ("3. divider_column_index is not None",      report.divider_column_index is not None),
        ("4. len(divider_row_numbers) == 4",         len(report.divider_row_numbers) == 4),
        ("5. missing_known_headers == ()",           report.missing_known_headers == ()),
        ("6. extra_headers == ()",                   report.extra_headers == ()),
    ]
    for label, passed in gates:
        status = "PASS" if passed else "FAIL"
        print(f"  [{status}] {label}")

    print()
    print("Gate 7: STAKEHOLDER CONFIRMATION — must be confirmed manually.")
    print(
        "  Operations must confirm that no downstream consumer (frontend, "
        "reports, ops queries) needs to slice rows by section origin. "
        "This makes Decision 3 a one-way door (see Phase 12 TDD §0 Outstanding)."
    )
    print("=" * 60)


def main(argv: list[str]) -> int:
    if len(argv) != 2:
        print(f"Usage: python -m scripts._sniff_schd_layout <path-to-workbook.xlsx>")
        return 2
    path = Path(argv[1])
    report = sniff_schd(path)
    _print_report(report)
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
