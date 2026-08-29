from __future__ import annotations

from collections.abc import Iterable, Iterator, Mapping
from dataclasses import dataclass
from datetime import date, datetime
import logging
from typing import NamedTuple

log = logging.getLogger(__name__)

from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText

from .models import SheetKind

PRIMARY_SHEET_NAME: str = "SCHD"
FALLBACK_SHEET_NAME: str = "SHIPPED (AA)"

# Stage 3 flip: SCHD is the new default. resolve_sheet falls back to
# FALLBACK_SHEET_NAME when PRIMARY_SHEET_NAME is requested but absent.
SHEET_NAME: str = PRIMARY_SHEET_NAME

KNOWN_HEADERS: frozenset[str] = frozenset({
    "SHIPPED", "PCB NOTES", "KIT NOTES", "SCHEDULING NOTES",
    "LINE 1", "LINE 2", "LINE 3", "JOB", "QTY", "SHIP DATE",
    "PROG", "MFG NOTES", "SMT LINES", "SMT PLCMNTS",
    "SHIP METHOD", "CUSTOMER", "SALES P", "DOC REL", "KIT REL",
    "CODE", "BOM COMPARE / PHOTOS",
})

MARKDOWN_HEADERS: frozenset[str] = frozenset({
    "MFG NOTES",
    "PCB NOTES",
    "SCHEDULING NOTES",
    "KIT NOTES",
})

DATA_BEARING_HEADERS: frozenset[str] = KNOWN_HEADERS - MARKDOWN_HEADERS

_BOLD = "**"
_ITALIC = "*"
_STRIKE = "~~"


# ---------------------------------------------------------------------------
# Header normalization (Phase 22 Part 1)
# ---------------------------------------------------------------------------

# 0-based position of a column within a sheet row, as produced by enumerate()
# over the header cells. A plain alias, not a newtype: it indexes openpyxl row
# tuples directly and gains nothing from wrapping.
ColumnIndex = int


def normalize_header(raw_cell_value: object) -> str | None:
    """Return the canonical form of a worksheet header cell value.

    Pre:   raw_cell_value is an openpyxl cell value of any type, including None.
    Post:  str inputs return stripped, or None when stripping empties them.
           Every non-str input returns None.  Pure; total.
           Mapping non-str to None is behaviour-preserving: a non-str header can
           satisfy neither `in KNOWN_HEADERS` nor `== divider_header` today.
    Raises: never.
    """
    if not isinstance(raw_cell_value, str):
        return None
    stripped = raw_cell_value.strip()
    return stripped if stripped else None


@dataclass(frozen=True)
class HeaderCollision:
    """One ignored duplicate occurrence of an already-bound header.

    Pre:  header is a normalized KNOWN_HEADERS member already bound to
          kept_index by resolve_column_indices.
    Post: immutable; ignored_index names the occurrence whose data is discarded.
    """

    header: str
    kept_index: ColumnIndex
    ignored_index: ColumnIndex


@dataclass(frozen=True)
class HeaderReconciliation:
    """How one sheet's header row lines up with KNOWN_HEADERS.

    Pre:  built by reconcile_headers from a normalized header row.
    Post: immutable; the four fields partition the comparison completely, so a
          consumer never has to re-derive one from the others.
    """

    sheet_name: str
    matched: frozenset[str]
    missing: frozenset[str]
    unrecognised: tuple[str, ...]
    collisions: tuple[HeaderCollision, ...]


def resolve_column_indices(
    normalized_headers: list[str | None],
    divider_header: str | None,
) -> Mapping[str, ColumnIndex]:
    """Map each recognised header to the 0-based column index carrying it.

    Pre:   normalized_headers is the output of normalize_header over the
           sheet's header row.  divider_header is layout.divider_header or None.
    Post:  returns one entry per KNOWN_HEADERS member present in the sheet,
           bound to its FIRST occurrence.  Later occurrences are reported in
           HeaderReconciliation.collisions and are never read.
           divider_header is excluded from the mapping — it is a sentinel, not
           a data column, and must not reach the extracted cell dict.
           Insertion order is sheet order.
    Raises: never.
    """
    column_indices: dict[str, ColumnIndex] = {}
    for index, header in enumerate(normalized_headers):
        if header is None or header == divider_header:
            continue
        if header not in KNOWN_HEADERS:
            continue
        if header in column_indices:
            continue
        column_indices[header] = index
    return column_indices


def reconcile_headers(
    normalized_headers: list[str | None],
    divider_header: str | None,
    sheet_name: str,
) -> HeaderReconciliation:
    """Report how the sheet's header row lines up with KNOWN_HEADERS.

    Pre:   normalized_headers is the output of normalize_header over the header
           row.  divider_header is layout.divider_header or None.
    Post:  pure; reads no module state beyond KNOWN_HEADERS.
           matched | missing == KNOWN_HEADERS and matched & missing == empty.
           unrecognised excludes divider_header and every None cell, and
           preserves sheet order.
           collisions carries one HeaderCollision per extra occurrence of an
           already-bound header — two extra occurrences yield two entries.
    Raises: never.
    """
    bound: dict[str, ColumnIndex] = {}
    unrecognised: list[str] = []
    collisions: list[HeaderCollision] = []

    for index, header in enumerate(normalized_headers):
        if header is None or header == divider_header:
            continue
        if header not in KNOWN_HEADERS:
            unrecognised.append(header)
            continue
        kept_index = bound.get(header)
        if kept_index is None:
            bound[header] = index
        else:
            collisions.append(
                HeaderCollision(
                    header=header, kept_index=kept_index, ignored_index=index
                )
            )

    matched = frozenset(bound)
    return HeaderReconciliation(
        sheet_name=sheet_name,
        matched=matched,
        missing=KNOWN_HEADERS - matched,
        unrecognised=tuple(unrecognised),
        collisions=tuple(collisions),
    )


def assert_constants_are_normal(extra_headers: Iterable[str] = ()) -> None:
    """Guard the header constants against drifting out of normal form.

    Pre:   extra_headers carries header constants declared in other modules
           that are compared against sheet headers — ingest._COLUMN_MAP keys.
           The parameter exists so the check runs without reader importing
           ingest; the caller owns the reverse direction, which already exists.
    Post:  returns normally iff every KNOWN_HEADERS member and every member of
           extra_headers equals its own normalize_header output.
           SheetLayout.divider_header is NOT checked here — it is covered at
           construction, which is strictly broader.
    Raises: AssertionError
    """
    for header in KNOWN_HEADERS:
        assert normalize_header(header) == header, (
            f"KNOWN_HEADERS member {header!r} is not in normal form"
        )
    for header in extra_headers:
        assert normalize_header(header) == header, (
            f"Header constant {header!r} is not in normal form"
        )


def _log_reconciliation(reconciliation: HeaderReconciliation) -> None:
    """Emit the per-read_rows header reconciliation at the levels of §1.4.

    Pre:   reconciliation was built for the sheet read_rows is about to iterate.
    Post:  missing and collisions log at WARNING, unrecognised at INFO, matched
           is not logged.  Emitted once per read_rows call, never per row.
    Raises: never.
    """
    if reconciliation.missing:
        log.warning(
            "reader.headers.missing sheet=%r headers=%s",
            reconciliation.sheet_name,
            sorted(reconciliation.missing),
        )
    for collision in reconciliation.collisions:
        log.warning(
            "reader.headers.collision sheet=%r header=%r kept_index=%d "
            "ignored_index=%d",
            reconciliation.sheet_name,
            collision.header,
            collision.kept_index,
            collision.ignored_index,
        )
    if reconciliation.unrecognised:
        log.info(
            "reader.headers.unrecognised sheet=%r headers=%s",
            reconciliation.sheet_name,
            list(reconciliation.unrecognised),
        )


# ---------------------------------------------------------------------------
# Disposition primitives (Phase 21)
# ---------------------------------------------------------------------------

def is_blank_cell(cell_value: str | None) -> bool:
    """Report whether an extracted cell carries no operator content.

    Pre:   cell_value is post-extraction — the output of cell_to_text or
           cell_to_markdown, hence str | None.
    Post:  True iff cell_value is None, the empty string, or a string whose
           characters are all whitespace. Pure; total.
    Raises: never.
    """
    if cell_value is None:
        return True
    return cell_value.strip() == ""

import enum

class RowDisposition(enum.Enum):
    """Enumerates why a sheet row was kept or dropped. Shared logging vocabulary
    for both decision sites: the sentinel test in _iter_data_rows and the
    post-extraction test in classify_row.

      divider_marked — DIVIDER sentinel cell was non-empty. Decided by
                       _is_divider_row in _iter_data_rows, BEFORE extraction.
                       classify_row never returns this member.
      blank          — no KNOWN_HEADERS column carried content.
      non_data       — only notes columns carried content; no job identity.
      data           — at least one data-bearing column carried content.
    """
    divider_marked = "divider_marked"
    blank = "blank"
    non_data = "non_data"
    data = "data"

def classify_row(extracted_cells: dict[str, str | None]) -> RowDisposition:
    """Return the disposition of an already-extracted SCHD / SHIPPED (AA) row.

    Pre:   extracted_cells maps KNOWN_HEADERS names to their post-extraction
           values. The caller has ALREADY applied the DIVIDER sentinel test and
           has not called this function for a marked row.
    Post:  returns exactly one of blank | non_data | data. Never returns
           divider_marked — that disposition is unreachable here by contract,
           because a marked row is skipped before extraction ever runs.
           Emptiness is decided by is_blank_cell throughout.
           Pure — mutates nothing, reads no module state beyond the two header
           frozensets.
    Raises: never.
    """
    if all(is_blank_cell(v) for v in extracted_cells.values()):
        return RowDisposition.blank
    
    if all(is_blank_cell(extracted_cells.get(k)) for k in DATA_BEARING_HEADERS):
        return RowDisposition.non_data
        
    return RowDisposition.data


# ---------------------------------------------------------------------------
# Sheet layout primitives (Stage 1)
# ---------------------------------------------------------------------------


class MissingDividerColumnError(Exception):
    """Raised when a SheetLayout declares divider_header but the resolved
    sheet's header row does not contain that column.

    Pre:  sheet_name is the actual sheet selected by resolve_sheet.
          expected_header is the layout's divider_header.
    Post: instance carries both for error formatting; no further processing.
    """

    def __init__(self, sheet_name: str, expected_header: str) -> None:
        self.sheet_name = sheet_name
        self.expected_header = expected_header
        super().__init__(
            f"Sheet {sheet_name!r} is missing required column "
            f"{expected_header!r}; cannot detect section dividers."
        )


class EmptySheetError(Exception):
    """Raised when the resolved sheet has fewer rows than layout.header_row.

    Pre:  sheet_name is the resolved sheet. expected_header_row is
          layout.header_row. actual_max_row is ws.max_row at validation time.
    Post: instance carries all three for diagnostics.
    """

    def __init__(
        self, sheet_name: str, expected_header_row: int, actual_max_row: int
    ) -> None:
        self.sheet_name = sheet_name
        self.expected_header_row = expected_header_row
        self.actual_max_row = actual_max_row
        super().__init__(
            f"Sheet {sheet_name!r} has {actual_max_row} rows; "
            f"layout requires header at row {expected_header_row}."
        )


@dataclass(frozen=True)
class SheetLayout:
    """Per-sheet structural rules consumed by read_rows.

    Pre:  header_row is 1-based and points at the row containing column
          names. divider_header, when set, names a column whose non-empty
          cell marks the row as a section-divider header to be skipped.
    Post: instances are immutable; equality is structural.

    Note: the sheet's name is the dict key in _LAYOUTS, not a field on
    this dataclass — there is exactly one source of truth for "what is
    this layout called."
    """

    header_row: int
    divider_header: str | None  # None = no divider filtering for this shape

    def __post_init__(self) -> None:
        """Reject a SheetLayout whose divider_header is not in normal form.

        Pre:   invoked by the dataclass machinery on every SheetLayout
               construction.
        Post:  returns normally iff divider_header is None or equals
               normalize_header(divider_header).  Applies to EVERY instance,
               whether or not it is registered in _LAYOUTS — a collection walk
               would miss a layout constructed elsewhere and silently re-open
               the two-vocabulary failure class.
        Raises: AssertionError
        """
        assert (
            self.divider_header is None
            or normalize_header(self.divider_header) == self.divider_header
        ), f"SheetLayout.divider_header {self.divider_header!r} is not in normal form"


class _LayoutEntry(NamedTuple):
    """Pairs the structural layout rules with the semantic kind for a sheet."""

    layout: SheetLayout
    kind: SheetKind


_LAYOUTS: dict[str, _LayoutEntry] = {
    "SCHD":         _LayoutEntry(SheetLayout(header_row=2, divider_header="DIVIDER"), SheetKind.live),
    "SHIPPED (AA)": _LayoutEntry(SheetLayout(header_row=1, divider_header=None),     SheetKind.historical),
}


def resolve_sheet(wb, requested: str) -> str:
    """Return the actual sheet name to load from *wb*.

    Pre:  wb is an opened openpyxl Workbook. requested is the caller's
          sheet name.
    Post: returns requested verbatim when present in wb.sheetnames.
          Falls back to FALLBACK_SHEET_NAME only when requested equals
          PRIMARY_SHEET_NAME and FALLBACK_SHEET_NAME is present. All
          other absent requests raise KeyError — no silent re-routing.
    Raises: KeyError when no acceptable sheet is found.
    """
    if requested in wb.sheetnames:
        return requested
    if requested == PRIMARY_SHEET_NAME and FALLBACK_SHEET_NAME in wb.sheetnames:
        return FALLBACK_SHEET_NAME
    raise KeyError(requested)


def resolve_layout(sheet_name: str) -> SheetLayout:
    """Return the registered SheetLayout for *sheet_name*.

    Pre:  sheet_name is the actual sheet name returned by resolve_sheet.
    Post: returns the registered SheetLayout.
    Raises: KeyError when no layout is registered for sheet_name — protects
            against silent shape drift when a new sheet is wired through
            resolve_sheet without a matching layout entry.
    """
    return _LAYOUTS[sheet_name].layout


def classify_sheet(sheet_name: str) -> SheetKind:
    """Return the SheetKind for *sheet_name*.

    Pre:  sheet_name is a string.
    Post: Returns SheetKind.live for PRIMARY_SHEET_NAME,
          SheetKind.historical for FALLBACK_SHEET_NAME.
    Raises: KeyError on any sheet name absent from _LAYOUTS — this is a drift
            assertion, not a runtime user-visible failure.  resolve_sheet only
            emits names present in _LAYOUTS, so under correct configuration
            the raise is unreachable.
    """
    return _LAYOUTS[sheet_name].kind


def cell_to_text(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, CellRichText):
        return "".join(getattr(run, "text", None) or str(run) for run in value)
    return str(value)


def cell_to_markdown(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if not isinstance(value, CellRichText):
        return str(value)
    parts: list[str] = []
    for run in value:
        text = getattr(run, "text", None) or str(run)
        font = getattr(run, "font", None)
        if font is not None:
            if getattr(font, "strike", False):
                text = f"{_STRIKE}{text}{_STRIKE}"
            if getattr(font, "i", False):
                text = f"{_ITALIC}{text}{_ITALIC}"
            if getattr(font, "b", False):
                text = f"{_BOLD}{text}{_BOLD}"
        parts.append(text)
    return "".join(parts)


def read_rows(
    path: str, sheet: str = SHEET_NAME
) -> Iterator[tuple[int, dict[str, str | None]]]:
    """Return an iterator over (row_number, cells) for every data row.

    Pre:  path is a readable .xlsx; sheet is either the default
          (SHEET_NAME == PRIMARY_SHEET_NAME) or an explicit name.
    Post: yields (row_number, cells) for each non-divider, non-empty row.
          Title rows (above layout.header_row) are skipped. Section-divider
          rows are skipped when the layout declares a divider_header.
          All-empty rows (including whitespace-only) and rows carrying only
          notes columns are skipped. DIVIDER is never included in cells.
          Header cells are normalized (stripped; non-str and empty-after-strip
          become None) before both divider and column resolution, and a header
          reconciliation is logged once per call — never per row.
    Raises (on CALL, not on first next()):
          KeyError                  — no acceptable sheet found.
          MissingDividerColumnError — layout requires DIVIDER column not present.
          EmptySheetError           — sheet has fewer rows than layout.header_row.
    Propagates: openpyxl load errors (FileNotFoundError, BadZipFile, ...).
    """
    wb = load_workbook(path, rich_text=True, data_only=True)
    sheet_name = resolve_sheet(wb, sheet)
    layout = resolve_layout(sheet_name)
    ws = wb[sheet_name]

    if ws.max_row is None or ws.max_row < layout.header_row:
        actual = ws.max_row or 0
        raise EmptySheetError(sheet_name, layout.header_row, actual)

    header_cells = next(
        ws.iter_rows(min_row=layout.header_row, max_row=layout.header_row)
    )
    # Normalization precedes BOTH divider resolution and column resolution.
    # Normalizing only inside the row loop would leave the divider lookup
    # matching raw text and the column lookup matching normalized text — two
    # vocabularies in one reader.
    normalized_headers = [normalize_header(c.value) for c in header_cells]
    _log_reconciliation(
        reconcile_headers(normalized_headers, layout.divider_header, sheet_name)
    )
    divider_index = _resolve_divider_index(layout, normalized_headers, sheet_name)
    column_indices = resolve_column_indices(normalized_headers, layout.divider_header)

    return _iter_data_rows(ws, column_indices, layout, divider_index)


def _resolve_divider_index(
    layout: SheetLayout, normalized_headers: list[str | None], sheet_name: str
) -> ColumnIndex | None:
    """Return the 0-based column index of the divider column, or None.

    Pre:  normalized_headers is the output of normalize_header over the resolved
          sheet's header row (may contain None entries for blank columns).
          layout.divider_header is already in normal form — SheetLayout asserts
          it at construction — so the comparison is normal-form on both sides.
    Post: returns the index when the layout requires divider filtering and
          the column is present. Returns None when the layout does not
          declare divider filtering.
    Raises: MissingDividerColumnError when the layout declares a divider
            header but headers does not contain it.
    """
    if layout.divider_header is None:
        return None
    try:
        return normalized_headers.index(layout.divider_header)
    except ValueError:
        raise MissingDividerColumnError(sheet_name, layout.divider_header)


def _iter_data_rows(
    ws,
    column_indices: Mapping[str, ColumnIndex],
    layout: SheetLayout,
    divider_index: ColumnIndex | None,
) -> Iterator[tuple[int, dict[str, str | None]]]:
    """Yield (row_number, cells) for each non-divider, non-empty data row.

    Pre:  ws is the resolved openpyxl Worksheet. column_indices is the output of
          resolve_column_indices — already filtered to KNOWN_HEADERS, already
          collision-resolved, divider column already excluded. layout is the
          resolved SheetLayout. divider_index is None or a valid 0-based index.
    Post: yields (row_number, cells) per row. Divider rows and all-empty rows
          are skipped. A recognised header whose index exceeds the row's width
          yields no key — classify_row reads an absent key as blank, identical
          to the pre-normalization behaviour for short rows.
    """
    for row_number, row in enumerate(
        ws.iter_rows(min_row=layout.header_row + 1, values_only=False),
        start=layout.header_row + 1,
    ):
        if _is_divider_row(row, divider_index):
            log.info(f"Row {row_number} skipped: disposition={RowDisposition.divider_marked.value}")
            continue
        cells: dict[str, str | None] = {}
        for header, index in column_indices.items():
            if index >= len(row):
                continue
            extract = cell_to_markdown if header in MARKDOWN_HEADERS else cell_to_text
            cells[header] = extract(row[index].value)
        
        disposition = classify_row(cells)
        if disposition == RowDisposition.blank:
            continue
        if disposition == RowDisposition.non_data:
            log.info(f"Row {row_number} skipped: disposition={disposition.value}")
            continue
            
        yield row_number, cells


def _is_divider_row(row, divider_index: int | None) -> bool:
    """Return True iff the row is a section-divider row.

    Pre:  row is a tuple of openpyxl Cell objects. divider_index is None or
          a valid 0-based column index.
    Post: True iff divider_index is not None AND the cell at that index
          exists AND the value satisfies Decision 5's non-empty rule:
            - None                          -> False
            - str, stripped == ""           -> False (whitespace-only typo)
            - str, stripped != ""           -> True
            - any non-None non-string value -> True
    """
    if divider_index is None:
        return False
    if divider_index >= len(row):
        return False
    value = row[divider_index].value
    if value is None:
        return False
    if isinstance(value, str):
        return value.strip() != ""
    return True

